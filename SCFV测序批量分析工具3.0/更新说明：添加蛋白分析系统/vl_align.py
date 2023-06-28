import os
import re
import time
from collections import Counter
from io import StringIO
from time import strftime

import openpyxl
from Bio import SeqIO, AlignIO
from Bio import pairwise2
from Bio.Align.Applications import MuscleCommandline
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord
from Bio.SeqUtils.ProtParam import ProteinAnalysis
from openpyxl.styles import Font

"""  参考序列
vl-λ:'SALTQPSALSVSLGQTARITCRGD/SLERYG/ANWYQQKPGQARVQVIY/GDD/IAPSGIPERFAGTKSGDTATLIISGAQAEDEADYYC/QMWDATGVGL/FGGGTHLTVLGQPKAAPS'
      -----------fr1---------/--cd1--/-------fr2------/cd2/----------------fr3-----------------/---cdr3---/--------fr4-------

vl-κ：'DVVLTQTPASLSLFPGESASISCKAS/QSLVHSDGKTY/LYWLLQKPGQSPQRLIY/QVS/SRDSGVPDRFTGSGSGTDFTLKISGVKAEDAGVYYC/AQATYYPRT/FGQGTKLEIKRADAKPS'
      -----------fr1------------/----cdr1---/-------fr2-------/cd2/----------------fr3-----------------/---cdr3--/-------fr4------- 
vh:'QVQLVESGGGLVQPGGSLRLSCAAS/GFTFERYA/MSWVRQAPGKGLEWVSL/ISRDGAS/YYSDSAKGRFTISRDNAKNTLYLQMDSLKPEDTAVYYC/ASASLAY/WGKGTLVTVSSASTKAPS'
    ----------fr1-----------/--cdr1--/-------fr2-------/--cd2--/-----------------fr3------------------/--cdr3-/-------fr4-------
"""


class InPut(object):  # 读取序列文件

    def __init__(self, __path):
        self.file_name = []
        self.__path = __path

    def listdir(self):
        # 读取.pro序列文件
        for __file in os.listdir(self.__path):
            if '.pro' in __file:
                __file_path = os.path.join(self.__path, __file)
                self.file_name.append(__file)

        # 读取.fasta文件
        # elif '.fasta' in __file:
        #     __file_path = os.path.join(self.__path, __file)
        #     print(__file_path)

        # print(self.file_name)
        return self.file_name, self.__path


class Vl_Align(object):

    def __init__(self, path, h_l):
        self.vl_l = Seq(
            'SALTQPSALSVSLGQTARITCRGDSLERYGANWYQQKPGQARVQVIYGDDIAPSGIPERFAGTKSGDTATLIISGAQAEDEADYYCQMWDATGVGLFGGGTHLTVLGQPKAAPS')
        self.vl_k = Seq(
            'DVVLTQTPASLSLFPGESASISCKASQSLVHSDGKTYLYWLLQKPGQSPQRLIYQVSSRDSGVPDRFTGSGSGTDFTLKISGVKAEDAGVYYCAQATYYPRTFGQGTKLEIKRADAKPS')
        self.vh = Seq(
            'QVQLVESGGGLVQPGGSLRLSCAASGFTFERYAMSWVRQAPGKGLEWVSLISRDGASYYSDSAKGRFTISRDNAKNTLYLQMDSLKPEDTAVYYCASASLAYWGKGTLVTVSSASTKAPS')
        self.h_l = h_l
        self.fil_path = path
        self.pro_dir = {}
        self.temp_list = []
        self.vl_count = 48
        self.dir_seq = {}
        self.total = 0
        self.last_time = 0
        self.font = Font(name='Times New Roman', size=11)
        self.wb = openpyxl.Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = 'All seq'
        self.ws2 = self.wb.create_sheet('CDR1', 1)
        self.ws3 = self.wb.create_sheet('CDR2', 2)
        self.ws4 = self.wb.create_sheet('CDR3', 3)
        self.ws5 = self.wb.create_sheet('Full-length Unique', 4)
        self.message = []

    # 读取蛋白文件
    def read_pro(self, name):
        if self.h_l==1:
            self.pro_dir.update({'refer': 'QVQLVESGGGLVQPGGSLRLSCAASGFTFERYAMSWVRQAPGKGLEWVSLISRDGASYYSDSAKGRFTISRDNAKNTLYLQMDSLKPEDTAVYYCASASLAYWGKGTLVTVSSASTKAPS'})
        else:
            self.pro_dir.update({'refer': 'DVVLTQTPASLSLFPGESASISCKASQSLVHSDGKTYLYWLLQKPGQSPQRLIYQVSSRDSGVPDRFTGSGSGTDFTLKISGVKAEDAGVYYCAQATYYPRTFGQGTKLEIKRADAKPS'})
        with open(self.fil_path + '\\' + name, 'r+') as f:
            pro = f.read()
        self.pro_dir.update({name[:-4]: str(pro)})

    # 蛋白分析
    def pro_analysis(self, pro):
        p = ProteinAnalysis(pro)
        length = p.length
        mw = p.molecular_weight()
        pI = p.isoelectric_point()
        ext1, ext2 = p.molar_extinction_coefficient()  # 消光系数（未形成二硫键，形成二硫键）
        Abs1 = ext1 / mw
        Abs2 = ext2 / mw
        return length, round(mw / 1000, 2), pI, ext1, ext2, Abs1, Abs2

    # 蛋白比对
    def align(self):
        # 定义一个方法，用来对self.part_vhh_list中的序列进行比对
        start_time = time.time()
        self.total = len(self.pro_dir)
        temp_name = f'temp{strftime("%Y%m%d%H%M%S")}'
        # 生成一个临时文件名，包含当前的时间戳
        n_align = self.total // self.vl_count  # 取整除 - 向下取接近商的整数
        # 计算需要比对的批次数，每批包含Vhh.n_batch个序列
        muscle_exe = 'muscle.exe'
        # 定义一个变量，存储外部程序的名称

        for i in self.pro_dir.keys():
            # 使用enumerate函数遍历序列列表，同时获取索引i和元素part
            self.temp_list.append(SeqRecord(seq=Seq(self.pro_dir[i]), id=str(i)))
        for n_count in range(n_align + 1):
            with open(f'{temp_name}.fasta', 'w') as f:
                # 使用with语句打开一个临时文件，模式为写入
                SeqIO.write(self.temp_list, f, 'fasta')
                # 使用SeqIO模块将新的列表写入临时文件，格式为fasta
            muscle_cline1 = MuscleCommandline(muscle_exe, input=f'{temp_name}.fasta')
            # 使用MuscleCommandline类创建一个命令行对象，传入外部程序名称和临时文件名作为参数
            try:
                stdout, stderr = muscle_cline1()
                # 调用命令行对象，执行外部程序，并获取标准输出和标准错误
            except:
                print('muscle进程报错！')
                # 如果出现异常，打印错误信息
                continue
            else:
                align_read = AlignIO.read(StringIO(stdout), 'fasta')
                # 如果没有异常，使用AlignIO模块从标准输出中读取比对结果，格式为fasta
                f1, c1, f2, c2, f3, c3, f4 = self.slicer(align_read)
                # # 调用slicer方法，传入比对结果，返回六个变量
                self.dictionary(f1, c1, f2, c2, f3, c3, f4, align_read)
                # 调用dictionary方法，传入六个变量和比对结果，进行后续处理
                if self.total > 48:
                    print(f'比对完成第{n_count + 1}批', '48 条序列')
                    self.message.append(f'比对完成第{n_count + 1}批'+'48 条序列')
                    self.total -= 48
                else:
                    print(f'比对完成第{n_count + 1}批', self.total, '条序列')
                    self.message.append(f'比对完成第{n_count + 1}批' + str(self.total) + '条序列')
                # 打印完成信息，包含当前批次的索引
            finally:
                os.remove(f'{temp_name}.fasta')

        end_time = time.time()
        # 记录结束时间
        self.last_time = end_time - start_time
        # 计算耗时，赋值给self.last_time
        print(f'比对耗时: {self.last_time} s')
        self.message.append(f'比对耗时: {self.last_time} s')
        # 打印耗时信息

    # 读取参考序列的fr1,cdr1,fr2,cdr2,fr3,cdr3,
    def slicer(self, refer):
        global seq
        if self.h_l==1:
            for j in refer:
                if j.id == 'refer':
                    seq = str(j.seq)
                    break
            f1 = re.search('[Q]-*[V]-*[Q]-*[L]', seq).span(0)[0]
            c1 = re.search('[G]-*[F]-*[T]-*[F]', seq).span(0)[0]
            f2 = re.search('[M]-*[S]-*[W]-*[V]', seq).span(0)[0]
            c2 = re.search('[I]-*[S]-*[R]-*[D]', seq).span(0)[0]
            f3 = re.search('[Y]-*[Y]-*[S]-*[D]', seq).span(0)[0]
            c3 = re.search('[A]-*[S]-*[A]-*[S]', seq).span(0)[0]
            f4 = re.search('[W]-*[G]-*[K]-*[G]', seq).span(0)[0]
            return f1, c1, f2, c2, f3, c3, f4
        else:
            for j in refer:
                if j.id == 'refer':
                    seq = str(j.seq)
                    break
            f1 = re.search('[D]-*[V]-*[V]-*[L]', seq).span(0)[0]
            c1 = re.search('[Q]-*[S]-*[L]-*[V]', seq).span(0)[0]
            f2 = re.search('[L]-*[Y]-*[W]-*[L]', seq).span(0)[0]
            c2 = re.search('[Q]-*[V]-*[S]', seq).span(0)[0]
            f3 = re.search('[S]-*[R]-*[D]-*[S]', seq).span(0)[0]
            c3 = re.search('[A]-*[Q]-*[A]-*[T]', seq).span(0)[0]
            f4 = re.search('[F]-*[G]-*[Q]-*[G]', seq).span(0)[0]
            return f1, c1, f2, c2, f3, c3, f4

    # 构建字典
    def dictionary(self, f1, c1, f2, c2, f3, c3, f4, align_read):
        """ 字典内容：ID = (全长序列,FR1,CDR1,FR2,CDR2,FR3,CDR3,FR4,Residues,Molecular weight (KDa)','pI','A280 Molar
        Extinction Coefficients (reduced)','A280 Extinction Coefficients 1mg/ml (reduced)', 'A280 Molar Extinction
        Coefficients (cystine bridges)','A280 Extinction Coefficients 1mg/ml (cystine bridges)',CDR123,cut_head) """

        for fc_seq in align_read:
            if fc_seq.id != 'refer':
                fr1 = str(fc_seq.seq)[f1:c1].replace('-', '')
                cdr1 = str(fc_seq.seq)[c1:f2].replace('-', '')
                fr2 = str(fc_seq.seq)[f2:c2].replace('-', '')
                cdr2 = str(fc_seq.seq)[c2:f3].replace('-', '')
                fr3 = str(fc_seq.seq)[f3:c3].replace('-', '')
                cdr3 = str(fc_seq.seq)[c3:f4].replace('-', '')
                fr4 = str(fc_seq.seq)[f4:-1].replace('-', '')
                full_length = str(fc_seq.seq).replace('-', '')
                cut_head = full_length[8:]
                cdr123 = cdr1 + '/' + cdr2 + '/' + cdr3
                length, mw, pI, ext1, ext2, Abs1, Abs2 = self.pro_analysis(full_length)
                self.dir_seq[fc_seq.id] = (
                    full_length, fr1, cdr1, fr2, cdr2, fr3, cdr3, fr4, length, mw, pI, ext1, Abs1,
                    ext2, Abs2, cdr123, cut_head)

    def all_seq(self):
        self.ws1.append(['ID', 'Complete seq', 'FR1', 'CDR1', 'FR2', 'CDR2', 'FR3', 'CDR3', 'FR4', 'Length',
                         'Molecular weight (KDa)',
                         'pI', 'A280 Molar Extinction Coefficients (reduced)',
                         'A280 Extinction Coefficients 1mg/ml (reduced)',
                         'A280 Molar Extinction Coefficients (cystine bridges)',
                         'A280 Extinction Coefficients 1mg/ml (cystine bridges)'])
        for k, v, in self.dir_seq.items():
            self.ws1.append(
                [k, v[0], v[1], v[2], v[3], v[4], v[5], v[6], v[7], v[8], v[9], v[10], v[11], v[12], v[13], v[14]])
            self.edit_font(self.ws1)

    def similar(self, n, t):
        global seq3
        threshold = t
        sequences = {}
        for id1, seq1 in self.dir_seq.items():
            is_similar = False
            for id2, seq2 in self.dir_seq.items():
                if id1 != id2:
                    seq3 = Seq(seq1[n])
                    seq4 = Seq(seq2[n])
                    score = pairwise2.align.globalxx(seq3, seq4, score_only=True)
                    similarity = score / max(len(seq3), len(seq4))
                    if similarity >= threshold:
                        if seq3 not in sequences:
                            sequences[str(seq3)] = []
                        sequences[str(seq3)].append(id1)
                        sequences[str(seq3)].append(id2)

                        if seq4 not in sequences:
                            sequences[str(seq4)] = []
                        sequences[str(seq4)].append(id1)
                        sequences[str(seq4)].append(id2)
                        is_similar = True

            if is_similar is False:
                if seq3 not in sequences:
                    sequences[str(seq3)] = []
                sequences[str(seq3)].append(id1)

        # 将相似和不相似的序列分组
        similar_seqs = {}
        dissimilar_seqs = {}
        for seq, ids in sequences.items():
            ids = list(set(ids))  # 去重
            if len(ids) > 1:
                similar_seqs[seq] = ids
            else:
                if seq not in dissimilar_seqs:
                    dissimilar_seqs[seq] = []
                dissimilar_seqs[seq].append(ids[0])

        # 将不相似的序列排在最后
        result = {}
        for seq, ids in similar_seqs.items():
            result[seq] = ids
        result.update(dissimilar_seqs)
        return result

    def cdr1(self):
        sim_cdr1 = []
        list_cdr1 = []
        self.ws2.cell(1, 1).value = ''
        cdr1_title = ['CDR1 seq', 'ID', 'Complete seq', 'FR1', 'CDR1', 'FR2', 'CDR2', 'FR3', 'CDR3', 'FR4',
                      'Length',
                      'Molecular weight (KDa)',
                      'pI', 'A280 Molar Extinction Coefficients (reduced)',
                      'A280 Extinction Coefficients 1mg/ml (reduced)',
                      'A280 Molar Extinction Coefficients (cystine bridges)',
                      'A280 Extinction Coefficients 1mg/ml (cystine bridges)']
        self.ws2.append(cdr1_title)  # 添加标题
        for s in self.dir_seq.values():  # 拉出CDR1
            list_cdr1.append(s[2])
        count_cdr1 = Counter(list_cdr1).most_common()  # 统计序列数
        for key in self.similar(2, 0.99).keys():
            sim_cdr1.append(key)
        self.ws2.cell(1, 1).value = 'Different CDR1：%d' % len(count_cdr1)  # 打印数据
        print('CDR1分类数：', len(count_cdr1))
        self.message.append('CDR1分类数：'+str(len(count_cdr1)))

        __rcdr1 = 3
        _row = 0
        for __c1 in sim_cdr1:
            self.ws2.cell(__rcdr1, 1).value = __c1
            _i = 1
            for _id, _value in self.dir_seq.items():
                _row = __rcdr1 + _i
                if __c1 == _value[2]:
                    self.ws2.cell(_row, 2).value = _id
                    self.ws2.cell(_row, 3).value = _value[0]
                    self.ws2.cell(_row, 4).value = _value[1]
                    self.ws2.cell(_row, 5).value = _value[2]
                    self.ws2.cell(_row, 6).value = _value[3]
                    self.ws2.cell(_row, 7).value = _value[4]
                    self.ws2.cell(_row, 8).value = _value[5]
                    self.ws2.cell(_row, 9).value = _value[6]
                    self.ws2.cell(_row, 10).value = _value[7]
                    self.ws2.cell(_row, 11).value = _value[8]
                    self.ws2.cell(_row, 12).value = _value[9]
                    self.ws2.cell(_row, 13).value = _value[10]
                    self.ws2.cell(_row, 14).value = _value[11]
                    self.ws2.cell(_row, 15).value = _value[12]
                    self.ws2.cell(_row, 16).value = _value[13]
                    self.ws2.cell(_row, 17).value = _value[14]
                    _i += 1
            __rcdr1 = _row + 1

        self.ws2.freeze_panes = 'C3'
        self.edit_font(self.ws2)

    def cdr2(self):
        sim_cdr2 = []
        list_cdr2 = []
        self.ws3.cell(1, 1).value = ''
        cdr2_title = ['CDR2 seq', 'ID', 'Complete seq', 'FR1', 'CDR1', 'FR2', 'CDR2', 'FR3', 'CDR3', 'FR4',
                      'Length',
                      'Molecular weight (KDa)',
                      'pI', 'A280 Molar Extinction Coefficients (reduced)',
                      'A280 Extinction Coefficients 1mg/ml (reduced)',
                      'A280 Molar Extinction Coefficients (cystine bridges)',
                      'A280 Extinction Coefficients 1mg/ml (cystine bridges)']
        self.ws3.append(cdr2_title)  # 添加标题
        for s in self.dir_seq.values():  # 拉出CDR1
            list_cdr2.append(s[4])
        count_cdr2 = Counter(list_cdr2).most_common()  # 统计序列数
        for key in self.similar(4, 0.99).keys():
            sim_cdr2.append(key)
        self.ws3.cell(1, 1).value = 'Different CDR2：%d' % len(count_cdr2)  # 打印数据
        print('CDR2分类数：', len(count_cdr2))
        self.message.append('CDR2分类数：'+str(len(count_cdr2)))

        __rcdr1 = 3
        _row = 0
        for __c1 in sim_cdr2:
            self.ws3.cell(__rcdr1, 1).value = __c1
            _i = 1
            for _id, _value in self.dir_seq.items():
                _row = __rcdr1 + _i
                if __c1 == _value[4]:
                    self.ws3.cell(_row, 2).value = _id
                    self.ws3.cell(_row, 3).value = _value[0]
                    self.ws3.cell(_row, 4).value = _value[1]
                    self.ws3.cell(_row, 5).value = _value[2]
                    self.ws3.cell(_row, 6).value = _value[3]
                    self.ws3.cell(_row, 7).value = _value[4]
                    self.ws3.cell(_row, 8).value = _value[5]
                    self.ws3.cell(_row, 9).value = _value[6]
                    self.ws3.cell(_row, 10).value = _value[7]
                    self.ws3.cell(_row, 11).value = _value[8]
                    self.ws3.cell(_row, 12).value = _value[9]
                    self.ws3.cell(_row, 13).value = _value[10]
                    self.ws3.cell(_row, 14).value = _value[11]
                    self.ws3.cell(_row, 15).value = _value[12]
                    self.ws3.cell(_row, 16).value = _value[13]
                    self.ws3.cell(_row, 17).value = _value[14]
                    _i += 1
            __rcdr1 = _row + 1

        self.ws3.freeze_panes = 'C3'
        self.edit_font(self.ws3)

    def cdr3(self):
        sim_cdr3 = []
        list_cdr3 = []
        self.ws4.cell(1, 1).value = ''
        cdr3_title = ['CDR3 seq', 'ID', 'Complete seq', 'FR1', 'CDR1', 'FR2', 'CDR2', 'FR3', 'CDR3', 'FR4',
                      'Length',
                      'Molecular weight (KDa)',
                      'pI', 'A280 Molar Extinction Coefficients (reduced)',
                      'A280 Extinction Coefficients 1mg/ml (reduced)',
                      'A280 Molar Extinction Coefficients (cystine bridges)',
                      'A280 Extinction Coefficients 1mg/ml (cystine bridges)']
        self.ws4.append(cdr3_title)  # 添加标题
        for s in self.dir_seq.values():  # 拉出CDR1
            list_cdr3.append(s[6])
        count_cdr3 = Counter(list_cdr3).most_common()  # 统计序列数
        for key in self.similar(6, 0.9).keys():
            sim_cdr3.append(key)
        self.ws4.cell(1, 1).value = 'Different CDR3：%d' % len(count_cdr3)  # 打印数据
        print('CDR3分类数：', len(count_cdr3))
        self.message.append('CDR3分类数：'+str(len(count_cdr3)))

        __rcdr1 = 3
        _row = 0
        for __c1 in sim_cdr3:
            self.ws4.cell(__rcdr1, 1).value = __c1
            _i = 1
            for _id, _value in self.dir_seq.items():
                _row = __rcdr1 + _i
                if __c1 == _value[6]:
                    self.ws4.cell(_row, 2).value = _id
                    self.ws4.cell(_row, 3).value = _value[0]
                    self.ws4.cell(_row, 4).value = _value[1]
                    self.ws4.cell(_row, 5).value = _value[2]
                    self.ws4.cell(_row, 6).value = _value[3]
                    self.ws4.cell(_row, 7).value = _value[4]
                    self.ws4.cell(_row, 8).value = _value[5]
                    self.ws4.cell(_row, 9).value = _value[6]
                    self.ws4.cell(_row, 10).value = _value[7]
                    self.ws4.cell(_row, 11).value = _value[8]
                    self.ws4.cell(_row, 12).value = _value[9]
                    self.ws4.cell(_row, 13).value = _value[10]
                    self.ws4.cell(_row, 14).value = _value[11]
                    self.ws4.cell(_row, 15).value = _value[12]
                    self.ws4.cell(_row, 16).value = _value[13]
                    self.ws4.cell(_row, 17).value = _value[14]
                    _i += 1
            __rcdr1 = _row + 1

        self.ws4.freeze_panes = 'C3'
        self.edit_font(self.ws4)

    def full_length_unique(self):
        f = 0
        full_seq = []
        th = 0.8
        full_dir = self.similar(16, th)
        for l in full_dir.values():
            if len(l) == 1:
                full_seq.append(l[0])
                f += 1
        print("全长独特序列数：%d" % f)
        self.message.append("全长独特序列数：%d" % f)

        self.ws5.cell(1, 1).value = 'full-length unique：%d' % f
        full_title = ['ID', 'Complete seq', 'FR1', 'CDR1', 'FR2', 'CDR2', 'FR3', 'CDR3', 'FR4',
                      'Length',
                      'Molecular weight (KDa)',
                      'pI', 'A280 Molar Extinction Coefficients (reduced)',
                      'A280 Extinction Coefficients 1mg/ml (reduced)',
                      'A280 Molar Extinction Coefficients (cystine bridges)',
                      'A280 Extinction Coefficients 1mg/ml (cystine bridges)']
        self.ws5.append(full_title)  # 添加标题
        _row = 2
        for __c1 in full_seq:
            _row += 1
            for _id, _value in self.dir_seq.items():
                if __c1 == _id:
                    self.ws5.cell(_row, 1).value = _id
                    self.ws5.cell(_row, 2).value = _value[0]
                    self.ws5.cell(_row, 3).value = _value[1]
                    self.ws5.cell(_row, 4).value = _value[2]
                    self.ws5.cell(_row, 5).value = _value[3]
                    self.ws5.cell(_row, 6).value = _value[4]
                    self.ws5.cell(_row, 7).value = _value[5]
                    self.ws5.cell(_row, 8).value = _value[6]
                    self.ws5.cell(_row, 9).value = _value[7]
                    self.ws5.cell(_row, 10).value = _value[8]
                    self.ws5.cell(_row, 11).value = _value[9]
                    self.ws5.cell(_row, 12).value = _value[10]
                    self.ws5.cell(_row, 13).value = _value[11]
                    self.ws5.cell(_row, 14).value = _value[12]
                    self.ws5.cell(_row, 15).value = _value[13]
                    self.ws5.cell(_row, 16).value = _value[14]

        self.ws5.freeze_panes = 'C3'
        self.edit_font(self.ws5)

    def edit_font(self, sheet):
        rows = sheet.max_row
        cols = sheet.max_column

        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                sheet.cell(r, c).font = self.font

    def save_file(self, output_path):
        stime = strftime('%Y%m%d%H%M%S')
        save_xlsx = os.path.join(output_path, 'Vl序列分析结果' + stime + '.xlsx')
        self.wb.save(save_xlsx)
        self.wb.close()

    def do(self):
        self.all_seq()
        self.cdr1()
        self.cdr2()
        self.cdr3()
        self.full_length_unique()



if __name__ == '__main__':
    path = input('请输入地址：')
    name, path = InPut(path).listdir()
    A = Vl_Align(path, 1)
    for i in name:
        A.read_pro(i)
    A.align()
    A.do()
    A.save_file(path)
